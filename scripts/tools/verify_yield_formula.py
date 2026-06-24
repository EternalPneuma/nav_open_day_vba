"""
验证 产品净值汇总_yyyymmdd.xlsx 中 30日年化收益率 是否符合单利公式。

对比新旧公式:
  旧(复利): ((V_t / V_base) ^ (365 / gap) - 1) * 100
  新(单利): (V_t / V_base - 1) * (365 / gap) * 100

仅验证 VBA 实际写入了计算值的行（审计列 = "计算填补" / "计算覆盖原始值"）。

用法:
    uv run python scripts/tools/verify_yield_formula.py
    uv run python scripts/tools/verify_yield_formula.py "产品净值汇总_20260611.xlsx"
"""

import sys
import os
from datetime import datetime
from openpyxl import load_workbook

# --- 配置 ---
DEFAULT_FILENAME = "产品净值汇总_20260611.xlsx"
TOLERANCE = 0.015  # 允许偏差 (百分点) — VBA Double 与 Python float 间微小舍入

# 列映射 (1-based)
COL_DATE = 1
COL_NAV = 4       # 净值
COL_YIELD = 7     # 30日年化收益率(%)
COL_AUDIT_TYPE = 11       # 处理方式
COL_AUDIT_BASE_DATE = 13  # 基准日期
COL_AUDIT_GAP = 14        # 实际间隔(天)

DATA_START_ROW = 2


def find_project_root() -> str:
    script_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.dirname(os.path.dirname(script_dir))


def build_nav_index(ws) -> dict:
    """构建 {date_long: nav} 索引。"""
    index: dict[int, float] = {}
    for r in range(DATA_START_ROW, ws.max_row + 1):
        d = ws.cell(r, COL_DATE).value
        n = ws.cell(r, COL_NAV).value
        if d is None or n is None or not isinstance(d, datetime):
            continue
        try:
            nav = float(n)
        except (ValueError, TypeError):
            continue
        if nav <= 0:
            continue
        date_long = int(d.timestamp() // 86400)
        # 同一日期保留第一条（最早出现）
        if date_long not in index:
            index[date_long] = nav
    return index


def calc_compound_yield(cur_nav: float, base_nav: float, gap: int) -> float:
    """复利年化: ((V_t / V_base) ^ (365 / gap) - 1) * 100"""
    return (pow(cur_nav / base_nav, 365.0 / gap) - 1.0) * 100.0


def calc_simple_yield(cur_nav: float, base_nav: float, gap: int) -> float:
    """单利年化: (V_t / V_base - 1) * (365 / gap) * 100"""
    return (cur_nav / base_nav - 1.0) * (365.0 / gap) * 100.0


def verify_sheet(ws, sheet_name: str) -> dict:
    """验证单个产品 sheet。"""
    nav_index = build_nav_index(ws)

    rpt = {
        "sheet": sheet_name,
        "vba_calc_rows": 0,       # VBA 实际写入计算值的行数
        "match_simple": 0,        # 匹配新单利公式
        "match_compound": 0,      # 匹配旧复利公式
        "match_both": 0,          # 两个公式结果接近（如收益率≈0）
        "match_neither": 0,       # 都不匹配 → 数据问题
        "neither_details": [],
    }

    for r in range(DATA_START_ROW, ws.max_row + 1):
        # 只验证 VBA 实际写入了计算值的行
        audit_type = str(ws.cell(r, COL_AUDIT_TYPE).value or "")
        is_vba_calc = ("填补" in audit_type or "覆盖" in audit_type) and ("保留" not in audit_type)
        if not is_vba_calc:
            continue

        rpt["vba_calc_rows"] += 1

        yield_val = ws.cell(r, COL_YIELD).value
        base_date = ws.cell(r, COL_AUDIT_BASE_DATE).value
        gap_val = ws.cell(r, COL_AUDIT_GAP).value
        cur_nav = ws.cell(r, COL_NAV).value
        cur_date = ws.cell(r, COL_DATE).value

        # 前提校验
        if any(v is None for v in [yield_val, base_date, gap_val, cur_nav]):
            rpt["match_neither"] += 1
            rpt["neither_details"].append(f"Row {r}: missing data (yield={yield_val}, base={base_date}, gap={gap_val}, nav={cur_nav})")
            continue
        if not isinstance(base_date, datetime):
            rpt["match_neither"] += 1
            rpt["neither_details"].append(f"Row {r}: base_date not datetime")
            continue

        try:
            yield_actual = float(yield_val)
            actual_gap = int(gap_val)
            cur_nav_f = float(cur_nav)
        except (ValueError, TypeError) as e:
            rpt["match_neither"] += 1
            rpt["neither_details"].append(f"Row {r}: conversion error: {e}")
            continue

        # 查找基准净值
        base_date_long = int(base_date.timestamp() // 86400)
        base_nav = nav_index.get(base_date_long)
        if base_nav is None:
            rpt["match_neither"] += 1
            rpt["neither_details"].append(
                f"Row {r}: base_nav not found for {base_date.date()}"
            )
            continue

        # 分别用两个公式计算
        simple_yield = calc_simple_yield(cur_nav_f, base_nav, actual_gap)
        compound_yield = calc_compound_yield(cur_nav_f, base_nav, actual_gap)

        match_s = abs(yield_actual - simple_yield) <= TOLERANCE
        match_c = abs(yield_actual - compound_yield) <= TOLERANCE

        if match_s and match_c:
            rpt["match_both"] += 1
        elif match_s:
            rpt["match_simple"] += 1
        elif match_c:
            rpt["match_compound"] += 1
        else:
            rpt["match_neither"] += 1
            rpt["neither_details"].append(
                f"Row {r}: date={cur_date.date() if isinstance(cur_date, datetime) else cur_date} | "
                f"cur_nav={cur_nav_f:.8f} | base_nav={base_nav:.8f} | gap={actual_gap} | "
                f"simple={simple_yield:.6f} | compound={compound_yield:.6f} | actual={yield_actual:.6f}"
            )

    return rpt


def main():
    root = find_project_root()
    filename = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_FILENAME
    filepath = os.path.join(root, filename)

    if not os.path.exists(filepath):
        print(f"[ERROR] File not found: {filepath}")
        sys.exit(1)

    print(f"File: {filepath}")
    print(f"New formula (simple) : (V_t/V_base - 1) * (365/gap) * 100")
    print(f"Old formula (compound): ((V_t/V_base) ^ (365/gap) - 1) * 100")
    print(f"Tolerance: {TOLERANCE} ppts")
    print(f"Only checking rows where VBA wrote calculated values (not 'retained original')")
    print("=" * 78)

    wb = load_workbook(filepath, data_only=True)

    all_results = []
    grand = {
        "vba_calc_rows": 0,
        "match_simple": 0,
        "match_compound": 0,
        "match_both": 0,
        "match_neither": 0,
    }

    for sn in wb.sheetnames:
        if sn == "数据摘要":
            continue
        ws = wb[sn]
        if ws.max_row < 2:
            continue

        r = verify_sheet(ws, sn)
        all_results.append(r)
        for k in grand:
            grand[k] += r[k]

        if r["vba_calc_rows"] > 0:
            # 判断该产品匹配哪个公式
            if r["match_simple"] > r["match_compound"]:
                verdict = "-> SIMPLE"
            elif r["match_compound"] > r["match_simple"]:
                verdict = "-> COMPOUND"
            else:
                verdict = "-> BOTH/EQUAL"
            print(f"  {verdict:<14s} | {r['sheet']:<30s} | "
                  f"calc={r['vba_calc_rows']:>4d} | "
                  f"simple={r['match_simple']:>4d} | "
                  f"compound={r['match_compound']:>4d} | "
                  f"both={r['match_both']:>4d} | "
                  f"neither={r['match_neither']:>4d}")

    wb.close()

    # 汇总
    print("=" * 78)
    print(f"TOTAL: VBA-calc-rows={grand['vba_calc_rows']} | "
          f"match-SIMPLE={grand['match_simple']} | "
          f"match-COMPOUND={grand['match_compound']} | "
          f"match-BOTH={grand['match_both']} | "
          f"match-NEITHER={grand['match_neither']}")

    # 判定结论
    print()
    if grand["match_simple"] > 0 and grand["match_compound"] == 0:
        print("[CONCLUSION] File was generated with NEW simple-interest formula. PASS.")
    elif grand["match_compound"] > 0 and grand["match_simple"] == 0:
        print("[CONCLUSION] File was generated with OLD compound-interest formula.")
        print("             VBA code has been updated; re-run Chart02_ExportProductSummary to regenerate.")
    elif grand["match_both"] == grand["vba_calc_rows"]:
        print("[CONCLUSION] All yields ~0 (both formulas agree). Cannot determine which was used.")
    elif grand["match_simple"] > 0 and grand["match_compound"] > 0:
        print(f"[CONCLUSION] Mixed — simple={grand['match_simple']}, compound={grand['match_compound']}.")
        print("             Some products may use different formulas. Check details.")
    else:
        print("[CONCLUSION] Indeterminate. See per-product breakdown above.")

    # 输出 neither 详情
    if grand["match_neither"] > 0:
        print(f"\n--- Neither-match details ({grand['match_neither']} rows) ---")
        for r in all_results:
            for detail in r["neither_details"]:
                print(f"  [{r['sheet']}] {detail}")

    return 0 if grand["match_neither"] == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
