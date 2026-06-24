"""诊断上层净值表与绘图净值表的主键匹配情况。

默认核对：
  - 上层产品净值数据(181)：B列日期、I列信托计划代码
  - 绘图净值数据：A列日期、B列信托计划代码
"""

from __future__ import annotations

import argparse
from collections import Counter
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


SOURCE_SHEET = "上层产品净值数据(181)"
TARGET_SHEET = "绘图净值数据"
EXCEL_EPOCH = datetime(1899, 12, 30)


def normalize_code(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def normalize_date(value: Any) -> str | None:
    """转换为 yyyy-mm-dd；覆盖 Excel 日期、序列值、yyyyMMdd 和常见文本。"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if 20_000_000 <= value <= 99_999_999 and float(value).is_integer():
            text = str(int(value))
            try:
                return date(int(text[:4]), int(text[4:6]), int(text[6:])).isoformat()
            except ValueError:
                return None
        try:
            return from_excel(value, epoch=EXCEL_EPOCH).date().isoformat()
        except (TypeError, ValueError, OverflowError):
            return None

    text = str(value).strip()
    if len(text) == 8 and text.isdigit():
        try:
            return date(int(text[:4]), int(text[4:6]), int(text[6:])).isoformat()
        except ValueError:
            return None
    for pattern in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日"):
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            pass
    return None


def describe_value(value: Any) -> str:
    return f"{type(value).__name__}: {value!r}"


def column_values(ws, column: int):
    return [row[0] for row in ws.iter_rows(min_row=2, min_col=column, max_col=column, values_only=True)]


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path, nargs="?", default=Path("上层产品净值数据库.xlsm"))
    args = parser.parse_args()

    wb = load_workbook(args.workbook, read_only=True, data_only=True, keep_vba=True)
    missing = [name for name in (SOURCE_SHEET, TARGET_SHEET) if name not in wb.sheetnames]
    if missing:
        raise SystemExit(f"缺少工作表：{', '.join(missing)}；实际工作表：{', '.join(wb.sheetnames)}")

    source = wb[SOURCE_SHEET]
    target = wb[TARGET_SHEET]
    source_dates, source_codes = column_values(source, 2), column_values(source, 9)
    target_dates, target_codes = column_values(target, 1), column_values(target, 2)

    source_keys = set()
    target_keys = set()
    for day, code in zip(source_dates, source_codes):
        day, code = normalize_date(day), normalize_code(code)
        if day and code:
            source_keys.add((day, code))
    for day, code in zip(target_dates, target_codes):
        day, code = normalize_date(day), normalize_code(code)
        if day and code:
            target_keys.add((day, code))

    source_date_set = {normalize_date(value) for value in source_dates} - {None}
    target_date_set = {normalize_date(value) for value in target_dates} - {None}
    source_code_set = {normalize_code(value) for value in source_codes} - {None}
    target_code_set = {normalize_code(value) for value in target_codes} - {None}

    def invalid_count(values, normalizer) -> int:
        return sum(normalizer(value) is None for value in values)

    print(f"工作簿: {args.workbook.resolve()}")
    print(f"{SOURCE_SHEET}: {len(source_dates)} 行；日期无效 {invalid_count(source_dates, normalize_date)}；代码无效 {invalid_count(source_codes, normalize_code)}")
    print(f"{TARGET_SHEET}: {len(target_dates)} 行；日期无效 {invalid_count(target_dates, normalize_date)}；代码无效 {invalid_count(target_codes, normalize_code)}")
    print()
    print(f"日期去重后：上层 {len(source_date_set)}，绘图 {len(target_date_set)}，交集 {len(source_date_set & target_date_set)}")
    print(f"代码去重后：上层 {len(source_code_set)}，绘图 {len(target_code_set)}，交集 {len(source_code_set & target_code_set)}")
    print(f"组合主键去重后：上层 {len(source_keys)}，绘图 {len(target_keys)}，交集 {len(source_keys & target_keys)}")
    print()
    print("绘图净值数据 A/B 列前 10 个非空原始值：")
    shown = 0
    for row in range(2, target.max_row + 1):
        raw_day = target.cell(row, 1).value
        raw_code = target.cell(row, 2).value
        if raw_day is not None or raw_code is not None:
            print(f"  第 {row} 行 | A: {describe_value(raw_day)} | B: {describe_value(raw_code)}")
            shown += 1
            if shown == 10:
                break

    if source_code_set & target_code_set:
        common_codes = sorted(source_code_set & target_code_set)[:10]
        print(f"\n代码交集样例：{common_codes}")
    if source_date_set & target_date_set:
        common_dates = sorted(source_date_set & target_date_set)[:10]
        print(f"日期交集样例：{common_dates}")


if __name__ == "__main__":
    main()
