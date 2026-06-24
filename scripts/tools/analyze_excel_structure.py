import sys
import os
import json
from datetime import datetime

import pandas as pd
import openpyxl
import xlrd


DEPENDENCIES = {
    "pandas": ">=1.0",
    "openpyxl": ">=3.0",
    "xlrd": ">=1.2",
}


def _parse_major_minor_patch(version_text: str):
    parts = []
    for p in str(version_text).split("."):
        try:
            parts.append(int(p))
        except Exception:
            break
    while len(parts) < 3:
        parts.append(0)
    return tuple(parts[:3])


def _require_python():
    if sys.version_info < (3, 8):
        raise RuntimeError(f"Python版本过低: {sys.version.split()[0]}，需要 >= 3.8")


def _check_dependency_versions():
    installed = {
        "pandas": getattr(pd, "__version__", "unknown"),
        "openpyxl": getattr(openpyxl, "__version__", "unknown"),
        "xlrd": getattr(xlrd, "__version__", "unknown"),
    }
    problems = []
    for name, constraint in DEPENDENCIES.items():
        need = constraint.replace(">=", "").strip()
        if need and installed.get(name) not in (None, "unknown"):
            if _parse_major_minor_patch(installed[name]) < _parse_major_minor_patch(need):
                problems.append(f"{name}=={installed[name]} (需要 {constraint})")
    if problems:
        raise RuntimeError("依赖版本不满足: " + ", ".join(problems))
    return installed


def _project_root():
    here = os.path.abspath(os.path.dirname(__file__))
    return os.path.abspath(os.path.join(here, "..", ".."))


def _normalize_column_name(name):
    if name is None or (isinstance(name, float) and pd.isna(name)):
        return ""
    s = str(name)
    s = s.replace("\u3000", " ").replace("\r", " ").replace("\n", " ")
    s = " ".join(s.split()).strip()
    return s


def _make_unique_columns(columns):
    seen = {}
    out = []
    for c in columns:
        base = _normalize_column_name(c)
        if base == "":
            base = "Unnamed"
        if base not in seen:
            seen[base] = 1
            out.append(base)
        else:
            seen[base] += 1
            out.append(f"{base}_{seen[base]}")
    return out


def _detect_header_row(file_path, sheet_name, max_scan_rows=30):
    df = pd.read_excel(
        file_path,
        sheet_name=sheet_name,
        header=None,
        engine="openpyxl",
        nrows=max_scan_rows,
    )
    best_row = 0
    best_score = -1.0
    for r in range(min(max_scan_rows, df.shape[0])):
        row = df.iloc[r]
        non_null = row.notna().sum()
        if non_null < 2:
            continue
        normalized = [_normalize_column_name(x) for x in row.tolist()]
        normalized = [x for x in normalized if x != ""]
        if len(normalized) == 0:
            continue
        unique_ratio = len(set(normalized)) / max(len(normalized), 1)
        score = float(non_null) + unique_ratio
        if score > best_score:
            best_score = score
            best_row = r
    return best_row


def analyze_excel_structure(input_file_relpath: str, output_json_relpath: str):
    root = _project_root()
    input_path = os.path.abspath(os.path.join(root, input_file_relpath))
    if not os.path.exists(input_path):
        raise FileNotFoundError(f"文件不存在: {input_path}")

    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=False)

    report = {
        "input_file": os.path.relpath(input_path, root),
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "python": sys.version.split()[0],
        "dependencies": _check_dependency_versions(),
        "sheet_names": list(wb.sheetnames),
        "sheets": [],
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = int(getattr(ws, "max_row", 0) or 0)
        max_col = int(getattr(ws, "max_column", 0) or 0)

        header_row = _detect_header_row(input_path, sheet_name=sheet_name)
        df = pd.read_excel(
            input_path,
            sheet_name=sheet_name,
            header=header_row,
            engine="openpyxl",
        )
        df.columns = _make_unique_columns(df.columns.tolist())

        df_effective = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
        dtypes = df_effective.dtypes.astype(str)
        dtype_dist = dtypes.value_counts().to_dict()
        null_ratio_by_col = df_effective.isna().mean().to_dict()
        total_cells = int(df_effective.shape[0] * df_effective.shape[1])
        overall_null_ratio = None
        if total_cells > 0:
            overall_null_ratio = float(df_effective.isna().sum().sum() / total_cells)

        report["sheets"].append(
            {
                "sheet_name": sheet_name,
                "excel_max_row": max_row,
                "excel_max_column": max_col,
                "pandas_header_row_0_based": int(header_row),
                "pandas_header_row_1_based": int(header_row) + 1,
                "data_shape": {"rows": int(df_effective.shape[0]), "cols": int(df_effective.shape[1])},
                "headers": list(df_effective.columns),
                "dtype_distribution": {str(k): int(v) for k, v in dtype_dist.items()},
                "null_ratio_overall": overall_null_ratio,
                "null_ratio_by_column": {str(k): float(v) for k, v in null_ratio_by_col.items()},
            }
        )

    out_path = os.path.abspath(os.path.join(root, output_json_relpath))
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    return out_path


def main():
    _require_python()
    root = _project_root()

    default_input = os.path.join("181-固收上层", "HS-181_多账套净值查询_20260220.xlsx")
    default_output = os.path.join("outputs", "analysis", "HS-181_多账套净值查询_20260220.structure.json")

    input_rel = sys.argv[1] if len(sys.argv) >= 2 else default_input
    output_rel = sys.argv[2] if len(sys.argv) >= 3 else default_output

    out_path = analyze_excel_structure(input_rel, output_rel)
    print(f"已生成JSON报告: {os.path.relpath(out_path, root)}")


if __name__ == "__main__":
    main()

